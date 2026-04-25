import os
import openpyxl
import mysql.connector
from dotenv import load_dotenv

load_dotenv()

db = mysql.connector.connect(
    host=os.environ.get("AWS_RDS_HOST"),
    user=os.environ.get("AWS_RDS_USER"),
    password=os.environ.get("AWS_RDS_PWD"),
    database=os.environ.get("AWS_RDS_DB")
)

cursor = db.cursor()

wb = openpyxl.load_workbook("attendings.xlsx")
ws = wb.active

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print("Columns found:", headers)

inserted = 0
skipped = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    first_name, last_name, username, password, role = row

    if not username:  # skip empty rows
        continue

    try:
        cursor.execute("""
            INSERT INTO users (first_name, last_name, username, password, role)
            VALUES (%s, %s, %s, %s, %s)
        """, (first_name, last_name, username, password, role))
        inserted += 1
    except mysql.connector.errors.IntegrityError as e:
        print(f"Skipped '{username}': {e}")
        skipped += 1

db.commit()
print(f"\nDone. {inserted} inserted, {skipped} skipped.")

cursor.close()
db.close()